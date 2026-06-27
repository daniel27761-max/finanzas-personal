# =============================================================================
# data_manager.py
# Responsabilidad única: toda la lógica de persistencia de datos.
# Lee y escribe en 'finanzas.xlsx' usando pandas + openpyxl.
# No importa Streamlit — este módulo es agnóstico a la UI.
# =============================================================================

import pandas as pd
from pathlib import Path
from datetime import date
from dateutil import parser as date_parser

# ---------------------------------------------------------------------------
# Constantes del dominio
# ---------------------------------------------------------------------------

ARCHIVO_EXCEL = "finanzas.xlsx"
HOJA_TRANSACCIONES = "Transacciones"
HOJA_PRESUPUESTOS = "Presupuestos"

# Categorías fijas de la aplicación. Cualquier cambio aquí se propaga
# automáticamente al desplegable de la UI.
CATEGORIAS = [
    "Alimentación",
    "Transporte",
    "Ocio",
    "Hogar",
    "Facturas",
    "Salud",
    "Ingresos",
]

# Tipos posibles de movimiento
TIPOS = ["Gasto", "Ingreso"]

# Esquema de columnas para la hoja de transacciones.
# Definirlo aquí evita errores de typo en el resto del código.
COLUMNAS_TX = ["Fecha", "Comercio", "Importe", "Tipo", "Categoría"]

# Esquema de columnas para la hoja de presupuestos.
COLUMNAS_PRESUPUESTO = ["Categoría", "Límite"]


# ---------------------------------------------------------------------------
# Funciones de inicialización y carga
# ---------------------------------------------------------------------------

def _crear_excel_vacio() -> None:
    """
    Crea el archivo Excel con las dos hojas necesarias y sus cabeceras.
    Se llama solo la primera vez, cuando el archivo no existe todavía.
    """
    with pd.ExcelWriter(ARCHIVO_EXCEL, engine="openpyxl") as writer:
        # Hoja de transacciones: empieza sin filas
        pd.DataFrame(columns=COLUMNAS_TX).to_excel(
            writer, sheet_name=HOJA_TRANSACCIONES, index=False
        )
        # Hoja de presupuestos: una fila por categoría, límite inicial 0
        df_pres = pd.DataFrame({
            "Categoría": CATEGORIAS,
            "Límite": [0.0] * len(CATEGORIAS),
        })
        df_pres.to_excel(writer, sheet_name=HOJA_PRESUPUESTOS, index=False)


def cargar_transacciones() -> pd.DataFrame:
    """
    Lee la hoja de Transacciones desde el Excel y devuelve un DataFrame.

    - Si el archivo no existe, lo crea y devuelve un DataFrame vacío.
    - Garantiza que la columna 'Fecha' es de tipo datetime.date para
      facilitar los filtros posteriores.

    Returns:
        pd.DataFrame con columnas: Fecha, Comercio, Importe, Tipo, Categoría
    """
    if not Path(ARCHIVO_EXCEL).exists():
        _crear_excel_vacio()
        return pd.DataFrame(columns=COLUMNAS_TX)

    df = pd.read_excel(
        ARCHIVO_EXCEL,
        sheet_name=HOJA_TRANSACCIONES,
        engine="openpyxl",
    )

    # Asegurar que todas las columnas esperadas existen aunque el archivo
    # haya sido creado manualmente por el usuario con alguna columna faltante.
    for col in COLUMNAS_TX:
        if col not in df.columns:
            df[col] = None

    # Normalizar la columna Fecha a objetos datetime.date puros
    if not df.empty:
        df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce").dt.date

    return df[COLUMNAS_TX]


def cargar_presupuestos() -> pd.DataFrame:
    """
    Lee la hoja de Presupuestos y devuelve un DataFrame.

    Si alguna categoría del catálogo no tiene fila en la hoja
    (p. ej. porque se añadió una categoría nueva), la añade con límite 0.

    Returns:
        pd.DataFrame con columnas: Categoría, Límite
    """
    if not Path(ARCHIVO_EXCEL).exists():
        _crear_excel_vacio()

    # Leer las hojas disponibles para manejar archivos creados con el nombre
    # incorrecto (bug previo: "Presupuesto" sin 's')
    import openpyxl
    wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
    hojas = wb.sheetnames
    wb.close()

    # Si la hoja correcta no existe pero existe la incorrecta, renombrarla
    if HOJA_PRESUPUESTOS not in hojas:
        if "Presupuesto" in hojas:
            wb = openpyxl.load_workbook(ARCHIVO_EXCEL)
            wb["Presupuesto"].title = HOJA_PRESUPUESTOS
            wb.save(ARCHIVO_EXCEL)
            wb.close()
        else:
            # No existe ninguna hoja de presupuestos: recrear el archivo
            _crear_excel_vacio()

    df = pd.read_excel(
        ARCHIVO_EXCEL,
        sheet_name=HOJA_PRESUPUESTOS,
        engine="openpyxl",
    )

    # Garantizar que todas las categorías actuales están presentes
    cats_existentes = set(df["Categoría"].tolist())
    filas_nuevas = [
        {"Categoría": cat, "Límite": 0.0}
        for cat in CATEGORIAS
        if cat not in cats_existentes
    ]
    if filas_nuevas:
        df = pd.concat([df, pd.DataFrame(filas_nuevas)], ignore_index=True)

    return df[COLUMNAS_PRESUPUESTO]


# ---------------------------------------------------------------------------
# Lógica de negocio: anti-duplicados
# ---------------------------------------------------------------------------

def es_duplicado(df: pd.DataFrame, fecha: date, comercio: str, importe: float) -> bool:
    """
    Comprueba si ya existe un registro con la misma fecha, comercio e importe.

    La comparación de texto normaliza espacios y mayúsculas para evitar
    falsos negativos por diferencias de formato.

    Args:
        df:       DataFrame de transacciones ya cargado.
        fecha:    Fecha de la transacción a verificar.
        comercio: Nombre del comercio/concepto.
        importe:  Importe exacto.

    Returns:
        True si el registro ya existe, False en caso contrario.
    """
    if df.empty:
        return False

    comercio_norm = comercio.strip().lower()

    mask = (
        (df["Fecha"] == fecha) &
        (df["Comercio"].str.strip().str.lower() == comercio_norm) &
        (df["Importe"].round(2) == round(importe, 2))
    )
    return mask.any()


# ---------------------------------------------------------------------------
# Escritura de datos
# ---------------------------------------------------------------------------

def _guardar_todo(df_tx: pd.DataFrame, df_pres: pd.DataFrame) -> None:
    """
    Función interna que sobreescribe el Excel completo con los datos actuales.

    Se usa openpyxl como motor para conservar el formato de celdas y
    mantener las dos hojas en el mismo archivo.

    Args:
        df_tx:   DataFrame completo de transacciones.
        df_pres: DataFrame completo de presupuestos.
    """
    with pd.ExcelWriter(ARCHIVO_EXCEL, engine="openpyxl") as writer:
        df_tx.to_excel(writer, sheet_name=HOJA_TRANSACCIONES, index=False)
        df_pres.to_excel(writer, sheet_name=HOJA_PRESUPUESTOS, index=False)


def añadir_transaccion(
    fecha: date,
    comercio: str,
    importe: float,
    tipo: str,
    categoria: str,
) -> tuple[bool, str]:
    """
    Añade una nueva transacción al Excel tras verificar duplicados.

    Flujo:
        1. Carga el estado actual del Excel.
        2. Ejecuta el filtro anti-duplicados.
        3. Si no hay duplicado, concatena la fila nueva y guarda.

    Args:
        fecha:     Fecha de la transacción (objeto date).
        comercio:  Nombre del comercio o concepto descriptivo.
        importe:   Importe en euros (float positivo).
        tipo:      "Gasto" o "Ingreso".
        categoria: Una de las categorías de CATEGORIAS.

    Returns:
        Tupla (éxito: bool, mensaje: str).
        El mensaje se muestra directamente en la UI.
    """
    df_tx = cargar_transacciones()
    df_pres = cargar_presupuestos()

    # --- Verificación anti-duplicados ---
    if es_duplicado(df_tx, fecha, comercio, importe):
        return False, (
            f"⚠️ Duplicado detectado: ya existe un registro del {fecha} "
            f"en '{comercio}' por {importe:.2f} €. Operación cancelada."
        )

    # --- Construir la nueva fila ---
    nueva_fila = pd.DataFrame([{
        "Fecha": fecha,
        "Comercio": comercio.strip(),
        "Importe": round(importe, 2),
        "Tipo": tipo,
        "Categoría": categoria,
    }])

    df_tx_actualizado = pd.concat([df_tx, nueva_fila], ignore_index=True)

    # --- Persistir ---
    _guardar_todo(df_tx_actualizado, df_pres)

    return True, f"✅ Registro añadido: {comercio} — {importe:.2f} € ({categoria})"


def actualizar_presupuestos(nuevos_limites: dict[str, float]) -> None:
    """
    Actualiza los límites de presupuesto para las categorías indicadas.

    Args:
        nuevos_limites: Diccionario {categoría: límite_en_euros}.
                        Solo se actualiza lo que se pasa; el resto no cambia.
    """
    df_tx = cargar_transacciones()
    df_pres = cargar_presupuestos()

    for cat, limite in nuevos_limites.items():
        df_pres.loc[df_pres["Categoría"] == cat, "Límite"] = round(limite, 2)

    _guardar_todo(df_tx, df_pres)


# ---------------------------------------------------------------------------
# Cálculos para el Dashboard
# ---------------------------------------------------------------------------

def calcular_gasto_por_categoria(df: pd.DataFrame, mes: int, año: int) -> pd.Series:
    """
    Suma los gastos agrupados por categoría para un mes y año concretos.

    Args:
        df:   DataFrame de transacciones.
        mes:  Número de mes (1-12).
        año:  Año de cuatro dígitos.

    Returns:
        pd.Series indexada por categoría con el gasto total acumulado.
        Las categorías sin gasto en ese mes no aparecen en la serie.
    """
    if df.empty:
        return pd.Series(dtype=float)

    mask = (
        (df["Tipo"] == "Gasto") &
        (pd.to_datetime(df["Fecha"]).dt.month == mes) &
        (pd.to_datetime(df["Fecha"]).dt.year == año)
    )
    return df[mask].groupby("Categoría")["Importe"].sum()


def calcular_resumen_mensual(df: pd.DataFrame) -> pd.DataFrame:
    """
    Agrega ingresos y gastos totales por mes para el gráfico comparativo.

    Returns:
        DataFrame con columnas: Mes (str "YYYY-MM"), Ingresos, Gastos.
        Ordenado cronológicamente.
    """
    if df.empty:
        return pd.DataFrame(columns=["Mes", "Ingresos", "Gastos"])

    df_copia = df.copy()
    df_copia["Fecha"] = pd.to_datetime(df_copia["Fecha"])
    df_copia["Mes"] = df_copia["Fecha"].dt.to_period("M").astype(str)

    resumen = df_copia.groupby(["Mes", "Tipo"])["Importe"].sum().unstack(fill_value=0)

    # Garantizar que ambas columnas existen aunque no haya registros de ese tipo
    for col in ["Gasto", "Ingreso"]:
        if col not in resumen.columns:
            resumen[col] = 0.0

    resumen = resumen.rename(columns={"Gasto": "Gastos", "Ingreso": "Ingresos"})
    resumen = resumen.reset_index().sort_values("Mes")

    return resumen[["Mes", "Ingresos", "Gastos"]]
